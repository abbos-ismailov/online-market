from openpyxl import load_workbook
from date import now_day, now_month, now_year
from sell import product_sell
from view import view_baza
from report import report
#### products bazani oldik
wb = load_workbook("./excel files/products_baza.xlsx")
sheet = wb.active
#### bosh excel ni oqidik
try:
    wsb = load_workbook(f"Add products_baza {now_day()}-{now_month()}-{now_year()}.xlsx")
except:
    wsb = load_workbook('./excel files/add_products.xlsx')
empty_sheet = wsb.active
empty_xlsx_m_row = empty_sheet.max_row + 1
products = []
def add_func():
    wsb.save(f"Add products_baza {now_day()}-{now_month()}-{now_year()}.xlsx")
    wb.save("./excel files/products_baza.xlsx")
    
def add_product():
    product_name = input("Mahsulotni nomini kiriting: ")
    product_count = input("Mahsulot soni: ")
    quantity_type = input("Qanday sotiladi (kg, dona, litr): ")
    product_quantity = input("Bir donasini miqdori yoki KG da sotilsa (1 kg) yoki LITR da sotilsa (1.5 litr, 1 litr): ")
    product_price = input("Mahsulot narxi: ")
    product_ish_sanasi = input("Mahsulotni ishlab chiqarilgan sanasi: ")
    product_saqlash_muddati = input("Mahsulotni saqlash muddati: ")
    all_price = int(product_count) * int(product_price)
    #### User kiritgan mahsulotlarni dict ga olib qoshdik
    product_dict = {
        "name": product_name.title(),
        "count": product_count,
        "quantity": product_quantity,
        "quantity_type": quantity_type,
        "price": product_price,
        "ish_sanasi": product_ish_sanasi,
        "saqlash_muddati": product_saqlash_muddati,
        "all_price": all_price,
    }
    products.append(product_dict)
    
    #### Soradik yana qoshiladimi yoki yetadimi
    yana_qosh = input("Yana qoshamizmi (1/0): ")
    if yana_qosh == "1":
        add_product()

    #### User kiritgan mahsulotlarni excel ga yozdik
    number = 0
    id = sheet.max_row + 1
    for i in range(id, len(products)+id):
        sheet[f"A{i}"].value = id - 1
        sheet[f"B{i}"].value = products[number]["name"]
        sheet[f"C{i}"].value = products[number]["count"]
        sheet[f"D{i}"].value = products[number]["quantity"]
        sheet[f"E{i}"].value = products[number]["price"]
        sheet[f"F{i}"].value = products[number]["ish_sanasi"]
        sheet[f"G{i}"].value = products[number]["saqlash_muddati"]
        sheet[f"H{i}"].value = products[number]["all_price"]
        number += 1
    number_2 = 0
    for i in range(empty_xlsx_m_row, len(products) + empty_xlsx_m_row):
        empty_sheet[f"A{i}"].value = empty_xlsx_m_row - 1
        empty_sheet[f"B{i}"].value = products[number_2]["name"]
        empty_sheet[f"C{i}"].value = products[number_2]["count"]
        empty_sheet[f"D{i}"].value = products[number_2]["quantity"]
        empty_sheet[f"E{i}"].value = products[number_2]["price"]
        empty_sheet[f"F{i}"].value = products[number_2]["ish_sanasi"]
        empty_sheet[f"G{i}"].value = products[number_2]["saqlash_muddati"]
        empty_sheet[f"H{i}"].value = products[number_2]["all_price"]
        number_2 += 1
def choice():
    savol = input("Mahsulot qo'shish (1): \nMahsulot sotish (2): \nMahsulotlarni korish (3): \nXisobot (4): \n >>> ")
    if savol == "1":
        add_product()
        add_func()
    elif savol == "2":
        product_sell()
    elif savol == "3":
        view_baza()
    elif savol == "4":
        report()
choice()
