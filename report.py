from openpyxl import load_workbook
from date import now_day, now_month, now_year
import qrcode

def report():
    sana_inp = input("Qaysi sanadagi sotilgan mahsulotlar kerak? ")
    try:
        wb_obj = load_workbook(f"Sell_products_baza {sana_inp}.xlsx") 
        sheet_obj = wb_obj.active
    except:
        print("Bunday sanada sotilmagan...")

    m_row = sheet_obj.max_row
    
    product_name_list = []
    product_count_list = []
    product_quantity_list = []
    ish_sanasi_list = []
    all_price_list = []

    for i in range(2, m_row + 1):
        product_name = sheet_obj.cell(row = i, column = 1)
        product_name_list.append(product_name.value)

        all_price = sheet_obj.cell(row = i, column = 2)
        all_price_list.append(all_price.value)

        product_count = sheet_obj.cell(row = i, column = 3)
        product_count_list.append(product_count.value)

        product_quantity = sheet_obj.cell(row = i, column = 4)
        product_quantity_list.append(product_quantity.value)
        
        ish_sanasi = sheet_obj.cell(row = i, column = 5)
        ish_sanasi_list.append(ish_sanasi.value)


    with open(f'Report {sana_inp}.doc', 'w') as f:
        for i in range(len(product_name_list)):
            text = f"""{product_name_list[i]}     {all_price_list[i]}     {product_count_list[i]}     {product_quantity_list[i]}     {ish_sanasi_list[i]}"""
            f.write(text)
            f.write('\n')
    
    data = f'https://github.com/abbos-ismailov/online-market/blob/master/Report%20{sana_inp}.doc'
    img = qrcode.make(data)
    img.save(f'Report {sana_inp}.png')