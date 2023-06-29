from openpyxl import load_workbook
from date import now_day, now_month, now_year

wb_obj = load_workbook("products_baza.xlsx") 
sheet_obj = wb_obj.active

new_obj = load_workbook("Sell_products_baza.xlsx")
new_sheet = new_obj.active

sell_wb = load_workbook(f"Sell_products_baza {now_day()}-{now_month()}-{now_year()}.xlsx")
sell_sheet = sell_wb.active
#### Products bazaniki bula
m_row = sheet_obj.max_row
max_col = sheet_obj.max_column

#### Sell products bazaniki
new_max_row = sell_sheet.max_row

sell_products_list = []
print(new_max_row)
def product_sell():
    product_name_list = []
    product_miq_list = []
    product_quantity_list = []
    all_price_list = []
    product_price_list = []
    
    
    #### Bu yerda qanday mahsulotlar borligini kordik
    #### Listga append qildik
    print("\n   Quyidagi mahsulotlar bor nimani sotmoqchisiz? \n\n")
    for i in range(2, m_row + 1):
        product_name = sheet_obj.cell(row = i, column = 2)
        product_name_list.append(product_name.value)

        product_count = sheet_obj.cell(row = i, column = 3)
        product_miq_list.append(product_count.value)
    
        product_quantity = sheet_obj.cell(row = i, column = 4)
        product_quantity_list.append(product_quantity.value)
        
        product_price = sheet_obj.cell(row = i, column = 5)
        product_price_list.append(product_price.value)
        
        all_price = sheet_obj.cell(row = i, column = 8)
        all_price_list.append(all_price.value)
        
        print(f"""        Mahsulot nomi: {product_name.value} 
        Jami soni: {product_count.value} 
        Narxi: {product_price.value} 
        Hajmi: {product_quantity.value}\n""")
    
    
    global sot_mah, sot_mah_bittasini_miq, sot_mah_bittasini_miq, sot_mah_miqdori        
    sot_mah = input("Sotiladigan mahsulot nomi: ")
    sot_mah_bittasini_miq = input("Sotiladigan mahsulotni sotiladigan hajmi: ")
    sot_mah_miqdori = int(input("Qancha sotmoqchisiz: "))
    
    #### Bu yerda hisob kitob ishlari olib borildi
    for i in range(len(product_name_list)):
        if sot_mah.title() == product_name_list[i] and int(sot_mah_miqdori) <= int(product_miq_list[i]) and sot_mah_bittasini_miq == str(product_quantity_list[i]):
            qolgan_mah_miq = int(product_miq_list[i]) - int(sot_mah_miqdori)
            
            all_price_list[i] = all_price_list[i] - (int(sot_mah_miqdori) * int(product_price_list[i]))
            product_miq_list[i] = qolgan_mah_miq
            print("In coming...")
            sell_product_dict = {
                "name": sot_mah,
                "quantity": sot_mah_bittasini_miq,
                "sot_mah_miqdori": sot_mah_miqdori,
                "qancha_pul": int(sot_mah_miqdori) * int(product_price_list[i]),
                "time": f"{now_day()}/{now_month()}/{now_year()}"
            }
            sell_products_list.append(sell_product_dict)
            break
        else:
            print("Elsega tushdi")
        
            
    #### Bu yerda ozgarishlarni yozib qoydik eski filega
    order = 0
    for i in range(2, len(product_miq_list) + 2):
        sheet_obj[f"C{i}"].value = product_miq_list[order]
        sheet_obj[f"H{i}"].value = all_price_list[order]
        order += 1
    
    #### Bu yerda yana soraldi yana sotiladimi yoki sotilmaydimi
    que_cont = input("Yana sotamizmi (1/0): ")
    if que_cont == "1":
        product_sell() 
    #### Yangi excel file ga yozyapmiz 
    number = 0
    try:
        id = new_max_row + 1
    except:
        id = 2
    print(id , "-> Bu id")
    for i in range(id, len(sell_products_list)+id):
        new_sheet[f"A{i}"].value = sell_products_list[number]["name"].title()
        new_sheet[f"B{i}"].value = sell_products_list[number]["qancha_pul"]
        new_sheet[f"C{i}"].value = sell_products_list[number]["sot_mah_miqdori"]
        new_sheet[f"D{i}"].value = sell_products_list[number]["quantity"]
        new_sheet[f"E{i}"].value = sell_products_list[number]["time"]
        number += 1
    #### Bu yerda excel fileni saqladik
    wb_obj.save(f"products_baza.xlsx")
    new_obj.save(f"Sell_products_baza {now_day()}-{now_month()}-{now_year()}.xlsx")