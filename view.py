from openpyxl import load_workbook
from date import now_day, now_month, now_year
wb = load_workbook("./excel files/products_baza.xlsx")
sheet_obj = wb.active

m_row = sheet_obj.max_row
max_col = sheet_obj.max_column


def view_baza():
    product_id_list = []
    product_name_list = []
    product_count_list = []
    product_quantity_list = []
    product_price_list = []
    ish_sanasi_list = []
    saqlash_muddati_list = []
    all_price_list = []
    
    for i in range(2, m_row + 1):
        product_id = sheet_obj.cell(row = i, column = 1)
        product_id_list.append(product_id.value)
        
        product_name = sheet_obj.cell(row = i, column = 2)
        product_name_list.append(product_name.value)

        product_count = sheet_obj.cell(row = i, column = 3)
        product_count_list.append(product_count.value)
    
        product_quantity = sheet_obj.cell(row = i, column = 4)
        product_quantity_list.append(product_quantity.value)
        
        product_price = sheet_obj.cell(row = i, column = 5)
        product_price_list.append(product_price.value)
        
        ish_sanasi = sheet_obj.cell(row = i, column = 6)
        ish_sanasi_list.append(ish_sanasi.value)
        
        saqlash_muddati = sheet_obj.cell(row = i, column = 7)
        saqlash_muddati_list.append(saqlash_muddati.value)
        
        all_price = sheet_obj.cell(row = i, column = 8)
        all_price_list.append(all_price.value)

    order = 0
    # id = m_row + 1
    for i in range(2, len(product_count_list) + 2):
        sheet_obj[f"A{i}"].value = product_id_list[order]
        sheet_obj[f"B{i}"].value = product_name_list[order]
        sheet_obj[f"C{i}"].value = product_count_list[order]
        sheet_obj[f"D{i}"].value = product_quantity_list[order]
        sheet_obj[f"E{i}"].value = product_price_list[order]
        sheet_obj[f"F{i}"].value = ish_sanasi_list[order]
        sheet_obj[f"G{i}"].value = saqlash_muddati_list[order]
        sheet_obj[f"H{i}"].value = all_price_list[order]
        order += 1

    wb.save(f"View Products Baza {now_day()}-{now_month()}-{now_year()}.xlsx")